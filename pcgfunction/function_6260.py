import pandas as pd
import numpy as np
# time

import time
from time import strftime
from datetime import datetime as datetimefunction
from datetime import timedelta
import os

import openpyxl
from openpyxl.styles import *
from openpyxl.utils import *
from openpyxl.drawing.image import Image

# import barcode
# from barcode.ean import EAN13
# from barcode.writer import ImageWriter
# from openpyxl.drawing.image import Image


# def outputbarcode(path, value, percent):
#    EAN = barcode.get_barcode_class('ean13')
#    # my_ean = EAN('000'+value, writer=ImageWriter())
#    my_ean = EAN13('000' + value, writer=ImageWriter())
#    fullname = my_ean.save(path+"/"+value, options={"write_text": False})
#    image = Image(fullname)
#    percent = percent
#    image.height = image.height * percent
#    image.width = image.width * percent
#    return image


def DN_print_6260(userfoldertimepath):
    dirkds = "./public/sample/MMKDS.xlsx"
    dfkds = pd.read_excel(dirkds, sheet_name=2, header=1, dtype=str)
    dfkds = dfkds[dfkds.columns[[3, 30, 5, 23, 8, 20]]].copy()
    dfkds.columns = ['fty', 'supplier',
                     'namezf', 'nameen', 'company', 'address']
    dfkds[dfkds.columns[5]] = dfkds[dfkds.columns[5]].astype(
        str).apply(lambda x: x.replace("\n", " ") if "\n" in x else x)

    df1039 = pd.read_excel(userfoldertimepath +
                           "/zrsd1039.xlsx", dtype=str).fillna('')
    df0002 = pd.read_excel(userfoldertimepath +
                           "/zrmm0002.xlsx", dtype=str).fillna('')
    df0002 = df0002[['物料', '物料說明', '型體代號+顏色代號', '型體顏色']].drop_duplicates(
        keep='first', inplace=False).reset_index(drop=True).copy()
    df1 = pd.merge(df1039, df0002, left_on=['物料'], right_on=['物料'], how='left', sort=False).drop_duplicates(
        keep='first', inplace=False).reset_index(drop=True).copy()
    df2 = df1.copy()
    df2 = df2.rename(columns={
        '廠別代號': 'Fty',
        '廠別名稱': 'FtyName',
        '客戶代號': 'Cus',
        '客戶名稱': 'CusName',
        '採購文件': 'STO',
        '項目': 'Item',
        '銷售文件': 'SO',
        '交貨': 'DN',
        '交貨狀態': 'Status',
        '物料': 'Article',
        '物料說明': 'Description',
        '型體代號+顏色代號': 'Article&Color',
        '型體顏色': 'Color',
        '原銷售數量': 'QtySTO',
        '出貨單數量': 'QtyDN',
        '基礎計量單位': 'Unit'
    })
    df3 = df2[['Fty', 'FtyName', 'Cus', 'CusName', 'STO', 'Item', 'SO', 'DN', 'Status',
               'Article', 'Description', 'Article&Color', 'Color', 'QtySTO', 'QtyDN', 'Unit']].copy()
    df3['DN'] = '0' + df3['DN']
    df3['GA'] = df3['Article'].str[:15]
    df3['Size'] = df3['Description'].apply(
        lambda x: x.split(",")[1] if "," in x else x).str.strip(' ')
    df3['Description'] = df3['Description'].apply(
        lambda x: x.split(",")[0] if "," in x else x)
    df3['QtySTO'] = df3['QtySTO'].astype(float)
    df3['QtyDN'] = df3['QtyDN'].astype(float)
    df3['QtyDN_A'] = 0
    df3.loc[df3['Status'] == 'A', 'QtyDN_A'] = df3['QtyDN']
    df3['QtyDN_C'] = 0
    df3.loc[df3['Status'] == 'C', 'QtyDN_C'] = df3['QtyDN']
    df3['QtyLack'] = df3['QtySTO'] - df3['QtyDN']
    df3['ItemQtySTO'] = df3.groupby(['STO', 'Item'])['QtySTO'].transform('sum')
    df3['ItemQtyDN_A'] = df3['QtyDN_A']
    df3['ItemQtyDN_C'] = df3.groupby(['STO', 'Item'])[
        'QtyDN_C'].transform('sum')
    df3['ItemQtyLack'] = df3.groupby(['STO', 'Item'])[
        'QtyLack'].transform('sum')
    df3['STO&SO&GA'] = df3['STO'] + df3['SO'] + df3['GA']
    dicItemcount = {}
    for stosoga in list(set(df3['STO&SO&GA'])):
        dicItemcount[stosoga] = len(
            df3[df3['STO&SO&GA'] == stosoga]['Item'].unique())
    df3['ItemCount'] = df3['STO&SO&GA'].apply(lambda x: dicItemcount[x])
    df3.loc[:, 'TotalQtySTO'] = 0
    df3.loc[:, 'TotalQtyDN'] = 0
    df3.loc[:, 'TotalQtyPGI'] = 0
    df3.loc[:, 'TotalQtyLack'] = 0
    cdt1 = df3['Status'] == 'A'
    dfSTOSOGADN = df3[cdt1][['STO', 'SO', 'GA', 'DN']].drop_duplicates(
        keep='first', inplace=False).reset_index(drop=True).copy()
    rawdatacol = ['STO&SO&GA', 'GA', 'ItemCount', 'Fty', 'FtyName', 'Cus', 'CusName', 'Description', 'Color', 'STO', 'Item', 'SO', 'Article', 'Size', 'Unit',
                  'ItemQtySTO', 'ItemQtyDN_C', 'ItemQtyLack',
                  'TotalQtySTO', 'TotalQtyDN', 'TotalQtyPGI', 'TotalQtyLack'
                  ]
    rawdata = df3[rawdatacol].drop_duplicates(
        keep='first', inplace=False).reset_index(drop=True).copy()
    dfdataall = pd.DataFrame(columns=rawdatacol)
    for i in range(len(dfSTOSOGADN)):
        cdtSTO = rawdata['STO'] == dfSTOSOGADN['STO'][i]
        cdtSO = rawdata['SO'] == dfSTOSOGADN['SO'][i]
        cdtGA = rawdata['GA'] == dfSTOSOGADN['GA'][i]
        cdtDN = df3['DN'] == dfSTOSOGADN['DN'][i]
        dfdataone = pd.merge(rawdata[cdtSTO & cdtSO & cdtGA],
                             df3[cdtDN][['STO', 'Item', 'ItemQtyDN_A']],
                             left_on=['STO', 'Item'],
                             right_on=['STO', 'Item'], how='left', sort=False)
        dfdataone = dfdataone.fillna(0).drop_duplicates(
            keep='first', inplace=False).reset_index(drop=True).copy()
        dfdataone['DN'] = dfSTOSOGADN['DN'][i]
        dfdataone['TotalQtySTO'] = dfdataone['ItemQtySTO'].sum()
        dfdataone['TotalQtyDN'] = dfdataone['ItemQtyDN_A'].sum()
        dfdataone['TotalQtyPGI'] = dfdataone['ItemQtyDN_C'].sum()
        dfdataone['TotalQtyLack'] = dfdataone['ItemQtyLack'].sum()
        dfdataall = pd.concat([dfdataall, dfdataone], ignore_index=True)
    dfdataall['Item'] = dfdataall['Item'].astype(int)
    dfdataall = dfdataall.sort_values(
        ['DN', 'STO', 'Item']).reset_index(drop=True)
    dfdataall['DN&STO&SO&GA'] = dfdataall['DN'] + \
        dfdataall['STO']+dfdataall['SO']+dfdataall['GA']
    printlistcol = ['DN&STO&SO&GA', 'ItemCount', 'DN', 'Fty', 'FtyName', 'Cus', 'CusName', 'Description', 'Color',
                    'TotalQtySTO', 'TotalQtyDN', 'TotalQtyPGI', 'TotalQtyLack']
    contentcolorigin = ['STO', 'Item', 'SO', 'Article', 'Size',
                        'ItemQtySTO', 'ItemQtyDN_A', 'ItemQtyDN_C', 'ItemQtyLack', 'Unit']
    contentcolnew = ['STO號', '項目', 'SO號', '物料',
                     'Size', 'STO數', 'DN數', '已出貨', '欠數', '單位']
    printlistdata = dfdataall[printlistcol].drop_duplicates(
        keep='first', inplace=False).reset_index(drop=True).copy()
    contentrow = 28
    printlistdata['BlankRow'] = contentrow - printlistdata['ItemCount']
    printlistdata['TitleRow'] = 1
    for row in range(1, len(printlistdata)):
        printlistdata.loc[row,
                          'TitleRow'] = printlistdata['TitleRow'][row-1]+6+contentrow
    printlistdata['TitleContentRow'] = printlistdata['TitleRow']+1
    printlistdata['ContentHeaderRow'] = printlistdata['TitleRow']+5
    printlistdata['ContentEndRow'] = printlistdata['ContentHeaderRow'] + \
        printlistdata['ItemCount']
    printlistdata['FinalRow'] = printlistdata['ContentHeaderRow']+contentrow
    dfconcat = pd.DataFrame(columns=contentcolorigin)
    for i in range(len(printlistdata)):
        dfheader = pd.DataFrame(
            np.full((6, len(contentcolorigin)), ''), columns=contentcolorigin)
        dfheader.iloc[0, 0] = dfkds.loc[dfkds['fty'] ==
                                        printlistdata['Fty'][i], 'company'].values[0]
        dfheader.iloc[1, 0] = "出貨方: "
        dfheader.iloc[2, 0] = "DN No.: "
        dfheader.iloc[3, 0] = "收貨方: "
        dfheader.iloc[4, 0] = "地址"
        dfheader.iloc[1, 4] = "出貨日: "
        dfheader.iloc[2, 4] = "說明: "
        dfheader.iloc[3, 4] = "顏色: "
        dfheader.iloc[5, :] = contentcolnew
        dfheader.iloc[1, 1] = printlistdata['Fty'][i] + \
            printlistdata['FtyName'][i]
        dfheader.iloc[2, 1] = printlistdata['DN'][i]
        dfheader.iloc[3, 1] = printlistdata['Cus'][i] + \
            printlistdata['CusName'][i]
        dfheader.iloc[4, 1] = dfkds.loc[dfkds['fty'] ==
                                        printlistdata['Cus'][i], 'address'].values[0]
        dfheader.iloc[1, 5] = strftime("%Y/%m/%d", time.localtime())
        dfheader.iloc[2, 5] = printlistdata['Description'][i]
        dfheader.iloc[3, 5] = printlistdata['Color'][i]
        dfcontent = dfdataall[dfdataall['DN&STO&SO&GA'] ==
                              printlistdata['DN&STO&SO&GA'][i]][contentcolorigin]
        dfblank = pd.DataFrame(np.full((printlistdata['BlankRow'][i], len(
            contentcolorigin)), ''), columns=contentcolorigin)
        dfblank.loc[0, 'Size'] = 'total: '
        dfblank.loc[0, 'ItemQtySTO'] = printlistdata['TotalQtySTO'][i]
        dfblank.loc[0, 'ItemQtyDN_A'] = printlistdata['TotalQtyDN'][i]
        dfblank.loc[0, 'ItemQtyDN_C'] = printlistdata['TotalQtyPGI'][i]
        dfblank.loc[0, 'ItemQtyLack'] = printlistdata['TotalQtyLack'][i]
        dfconcat = pd.concat([dfconcat, dfheader])
        dfconcat = pd.concat([dfconcat, dfcontent])
        dfconcat = pd.concat([dfconcat, dfblank])

    tm = (datetimefunction.now()+timedelta(hours=7)).strftime("%Y%m%d_%H%M%S")
    outputfilename = "DNprint_"+tm+".xlsx"
    dfconcat.to_excel(userfoldertimepath+"/"+outputfilename,
                      index=False, header=False)
    wb = openpyxl.load_workbook(userfoldertimepath+"/"+outputfilename)
    ws = wb[wb.sheetnames[0]]
    titlefont = Font(size=18.5, bold=True)
    alignmentcenter = Alignment(horizontal='center', vertical='center')
    alignmentright = Alignment(horizontal='right', vertical='center')
    fontborder = Side(border_style='medium', color='FF000000')
    colwidth = [10, 5, 10, 20, 15, 7, 7, 7, 7, 6]

    dfdnlist = printlistdata
    for index in dfdnlist.index:
        # title 合併儲存格
        ws.merge_cells(
            "A"+str(dfdnlist['TitleRow'][index])+":J"+str(dfdnlist['TitleRow'][index]))
        # title 改字體大小
        ws.cell(dfdnlist['TitleRow'][index], 1).font = Font(size=16, bold=True)
        # title 至中
        ws.cell(dfdnlist['TitleRow'][index], 1).alignment = alignmentcenter
        # title row height
        ws.row_dimensions[dfdnlist['TitleRow'][index]].height = 31
        # barcode
        # image = outputbarcode(userfoldertimepath, dfdnlist['DN號'][index], 0.2)
        # image.anchor = "D"+str(dfdnlist['TitleContentRow'][index])
        # ws.add_image(image)

    # 小列高
        for i in range(dfdnlist['TitleContentRow'][index], dfdnlist['FinalRow'][index]+1):
            ws.row_dimensions[i].height = 11

        for col in range(1, len(dfconcat.columns)+1):
            # columns 欄寬
            ws.column_dimensions[get_column_letter(
                col)].width = colwidth[col-1]
            # header 至中
            ws.cell(dfdnlist['ContentHeaderRow'][index],
                    col).alignment = alignmentcenter
            # content 字大小
            for row in range(dfdnlist['TitleContentRow'][index], dfdnlist['ContentEndRow'][index]+1+1):
                ws.cell(row, col).font = Font(size=9, bold=False)
            # content網線
            for row in range(dfdnlist['ContentHeaderRow'][index], dfdnlist['ContentEndRow'][index]+1):
                ws.cell(row, col).border = Border(top=fontborder,
                                                  bottom=fontborder, left=fontborder, right=fontborder)
                ws.cell(row, col).font = Font(size=9, bold=False)

        for col in range(6, 9+2):
            for row in range(dfdnlist['ContentHeaderRow'][index]+1, dfdnlist['ContentEndRow'][index]+1):
                ws.cell(row, col).alignment = alignmentright

    ws.print_options.horizontalCentered = True
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.set_printer_settings(paper_size=9, orientation='portrait')

    wb.save(userfoldertimepath+"/"+outputfilename)

    return outputfilename
